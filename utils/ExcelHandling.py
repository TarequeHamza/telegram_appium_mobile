import openpyxl


def open_and_read_excel_file(file, sheetName, column_value):
    filename = file
    workbook = openpyxl.load_workbook(filename, data_only=True)
    sheet = workbook[sheetName]
    column = int(column_value)
    m_row = sheet.max_row
    my_list = []  # created an empty list
    for i in range(2, m_row):  # Here I have started the loop from 2 as I want to skip the column heading value in
        # output
        cell_obj = sheet.cell(row=i, column=column)
        # print(cell_obj.value)
        if (cell_obj.value != None):
            my_list.append(cell_obj.value)
            continue
        else:
            break
    return my_list


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rowno = int(rowno)
    colno = int(colno)
    return sheet.cell(rowno, colno).value


def writeData(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rowno = int(rowno)
    colno = int(colno)
    sheet.cell(rowno, colno).value = str(data)
    workbook.save(file)
