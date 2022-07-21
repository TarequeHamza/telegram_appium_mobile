import openpyxl


def open_and_read_excel_file(file, sheetName, column_value):
    filename = file
    # wb_obj = openpyxl.load_workbook(filename, data_only=True)
    workbook = openpyxl.load_workbook(filename, data_only=True)
    # sheet_obj = wb_obj.get_sheet_by_name(sheetname)
    sheet = workbook[sheetName]
    column = int(column_value)
    m_row = sheet.max_row
    my_list = []  # created an empty list
    for i in range(2, m_row):  # Here I have started the loop from 2 as I want to skip the column heading value in
        # output
        cell_obj = sheet.cell(row=i, column=column)
        # print(cell_obj.value)
        if (cell_obj.value != None):
            my_list.append(cell_obj.value)
            continue
        else:
            break
    return my_list


# print(open_and_read_excel_file("Contract_Number_List.xlsx", "Sheet1", 2))


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


# print(getRowCount("Contract_Number_List.xlsx", "Sheet1"))


def getColCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rowno = int(rowno)
    colno = int(colno)
    return sheet.cell(rowno, colno).value


print(readData("Contract_Number_List.xlsx", "Sheet1", "2", "3"))


def writeData(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rowno = int(rowno)
    colno = int(colno)
    sheet.cell(rowno, colno).value = data
    workbook.save(file)


writeData("Contract_Number_List.xlsx", "Sheet1", "2", "3", "ok")


from Pages.BasePage import BasePage
from Pages.SendMessagePage import SendMessagePage
from Pages.ContactSavePage import SaveContactPage
from utils.ExcelHandling import *
import os.path


class SendMessageTest(BasePage):
    def test_SendMessage(self):
        send = SendMessagePage(self.driver)
        save = SaveContactPage(self.driver)
        excel = os.path.abspath('..\TestData\Contract_Number_List.xlsx')
        firstName = open_and_read_excel_file(excel, "Sheet1", 1)
        lastName = open_and_read_excel_file(excel, "Sheet1", 2)
        number = open_and_read_excel_file(excel, "Sheet1", 3)
        send.click_message_icon()
        for i in range(len(number)):
            send.click_search_icon()
            send.type_phone_number(number[i])
            Text = f"No results found for '{number[i]}'"
            invite = "INVITE"
            if Text == save.get_text():
                save.click_backArrow_icon()
                save.click_new_contact_button()
                save.type_first_name(firstName[i])
                save.type_last_name(lastName[i])
                save.type_phone_number(number[i])
                save.click_save_button()
                send.click_search_icon()
                send.type_phone_number(number[i])

                if 1==1: #unimportant

                    writeData(excel, "Sheet1", i + 2, 4, "1")
                    save.click_backArrow_icon()

                    send.click_expected_profile()
                    send.type_message(f"Hi {firstName[i]} {lastName[i]},"
                                      "\n\nNext Friday And Saturday Our Office Will Remain Closed, Enjoy Your Time!,"
                                      "\n\nHR, QUPS")
                    send.click_send_button()
                    send.click_back_button()
                    writeData(excel, "Sheet1", i + 2, 4, "Sent")

            elif invite == save.get_text_for_invite():
                writeData(excel, "Sheet1", i + 2, 4, "0")
                save.click_backArrow_icon()


            else:
                print("1111")
                writeData(excel, "Sheet1", i + 2, 4, "1")
                save.click_backArrow_icon()






